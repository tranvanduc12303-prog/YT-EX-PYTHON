import re
import openpyxl
from state import download_state, email_task_state, app_stats, DownloadCancelled

def yt_progress_hook(d):
    if download_state["stop_requested"]:
        raise DownloadCancelled("Download stopped by user")

def is_youtube(url: str) -> bool:
    return "youtube.com" in url.lower() or "youtu.be" in url.lower()

def extract_links_from_workbook(wb: openpyxl.Workbook) -> list[dict]:
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip()
                hyperlink = getattr(cell, "hyperlink", None)
                href = hyperlink.target if hyperlink and hyperlink.target else ""

                candidates = {val, href}
                for candidate in candidates:
                    if candidate and ("youtube" in candidate or "youtu.be" in candidate):
                        results.append({
                            "sheet": sheet_name,
                            "row": cell.row,
                            "col": cell.column,
                            "col_letter": cell.column_letter,
                            "title": val if val != candidate else "",
                            "url": candidate,
                            "valid": is_youtube(candidate),
                        })
                        break
    return results

def is_tiktok(url: str) -> bool:
    return "tiktok.com" in url.lower()

def extract_tiktok_links_from_workbook(wb: openpyxl.Workbook) -> list[dict]:
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip()
                hyperlink = getattr(cell, "hyperlink", None)
                href = hyperlink.target if hyperlink and hyperlink.target else ""

                candidates = {val, href}
                for candidate in candidates:
                    if candidate and (is_tiktok(candidate) or candidate.startswith("@")):
                        url = candidate
                        if candidate.startswith("@"):
                            url = f"https://www.tiktok.com/{candidate}"
                        
                        results.append({
                            "sheet": sheet_name,
                            "row": cell.row,
                            "col": cell.column,
                            "col_letter": cell.column_letter,
                            "title": val if val != candidate else "",
                            "url": url,
                            "valid": is_tiktok(url),
                        })
                        break
    return results

EMAIL_REGEX = re.compile(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+')

def sort_emails(emails):
    unique_emails = list(set([e.lower() for e in emails]))
    unique_emails.sort(key=lambda e: (e.split('@')[1], e.split('@')[0]))
    return unique_emails

def process_email_thread_func(input_path, output_path, source_col, target_col):
    email_task_state["status"] = "running"
    email_task_state["progress"] = 0
    email_task_state["processed_rows"] = 0
    email_task_state["emails_found"] = 0
    email_task_state["error_msg"] = ""
    email_task_state["sheet_stats"] = {}
    email_task_state["output_file"] = ""

    try:
        wb = openpyxl.load_workbook(input_path)
        
        # Calculate total rows first
        total_rows = 0
        for sheet_name in wb.sheetnames:
            total_rows += wb[sheet_name].max_row - 1 # exclude header
        email_task_state["total_rows"] = max(total_rows, 1)

        processed = 0
        total_found = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            email_task_state["sheet_stats"][sheet_name] = 0
            
            if ws.cell(row=1, column=target_col).value is None:
                ws.cell(row=1, column=target_col, value="Emails Được Trích Xuất")

            for row in range(2, ws.max_row + 1):
                if email_task_state["status"] == "stopped":
                    break

                cell_value = ws.cell(row=row, column=source_col).value
                if cell_value and isinstance(cell_value, str):
                    found_emails = EMAIL_REGEX.findall(cell_value)
                    if found_emails:
                        sorted_emails = sort_emails(found_emails)
                        ws.cell(row=row, column=target_col, value=", ".join(sorted_emails))
                        
                        count = len(sorted_emails)
                        total_found += count
                        email_task_state["sheet_stats"][sheet_name] += count
                        
                processed += 1
                email_task_state["processed_rows"] = processed
                email_task_state["emails_found"] = total_found
                email_task_state["progress"] = int((processed / email_task_state["total_rows"]) * 100)
            
            if email_task_state["status"] == "stopped":
                break

        wb.save(output_path)
        email_task_state["output_file"] = output_path
        app_stats["email_extracted_total"] += total_found

        if email_task_state["status"] != "stopped":
            email_task_state["status"] = "completed"
            email_task_state["progress"] = 100

    except Exception as e:
        email_task_state["status"] = "error"
        email_task_state["error_msg"] = str(e)
