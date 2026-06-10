import re
import argparse
import openpyxl
import os

# Biểu thức chính quy (Regex) để tìm email
EMAIL_REGEX = re.compile(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+')

def sort_emails(emails):
    """
    Loại bỏ email trùng lặp và sắp xếp có logic.
    Ưu tiên sắp xếp theo tên miền (domain) trước, sau đó tới tên người dùng (local part).
    """
    # Xoá trùng lặp và chuyển thành chữ thường để chuẩn hoá
    unique_emails = list(set([e.lower() for e in emails]))
    
    # Sắp xếp theo: Domain -> Local part
    unique_emails.sort(key=lambda e: (e.split('@')[1], e.split('@')[0]))
    return unique_emails

def process_excel(input_path, output_path, source_col=1, target_col=2):
    """
    Đọc file excel, trích xuất email từ cột nguồn, sắp xếp và ghi vào cột đích.
    """
    if not os.path.exists(input_path):
        print(f"❌ Lỗi: Không tìm thấy file '{input_path}'")
        return

    print(f"📂 Đang mở file: {input_path}")
    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"❌ Lỗi khi đọc file: {e}")
        return

    total_emails_found = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  ➜ Đang quét sheet: {sheet_name}...")
        
        # Cập nhật tiêu đề cho cột đích nếu đang trống
        if ws.cell(row=1, column=target_col).value is None:
            ws.cell(row=1, column=target_col, value="Emails Được Trích Xuất")

        # Quét từng hàng bắt đầu từ hàng 2 (bỏ qua tiêu đề)
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=source_col).value
            
            if cell_value and isinstance(cell_value, str):
                # Tìm tất cả email trong văn bản của ô
                found_emails = EMAIL_REGEX.findall(cell_value)
                
                if found_emails:
                    sorted_emails = sort_emails(found_emails)
                    # Ghi danh sách email đã sắp xếp, cách nhau bằng dấu phẩy
                    ws.cell(row=row, column=target_col, value=", ".join(sorted_emails))
                    total_emails_found += len(sorted_emails)

    try:
        wb.save(output_path)
        print(f"✅ Hoàn tất! Đã tìm thấy và sắp xếp tổng cộng {total_emails_found} email độc nhất.")
        print(f"💾 File kết quả được lưu tại: {output_path}")
    except PermissionError:
        print(f"❌ Lỗi: Không thể lưu file '{output_path}'. Hãy chắc chắn rằng bạn đã đóng file này trên Excel trước khi chạy tool.")
    except Exception as e:
        print(f"❌ Lỗi khi lưu file: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Tool trích xuất và sắp xếp Email từ file Excel.")
    parser.add_argument("input", help="Đường dẫn tới file Excel đầu vào (ví dụ: data.xlsx)")
    parser.add_argument("-o", "--output", help="Đường dẫn file đầu ra (Mặc định: tự động thêm '_emails' vào tên file gốc)", default=None)
    parser.add_argument("-c", "--col", type=int, help="Vị trí Cột chứa văn bản gốc cần quét (Mặc định: 1 - Cột A)", default=1)
    parser.add_argument("-t", "--target", type=int, help="Vị trí Cột để ghi email kết quả (Mặc định: 2 - Cột B)", default=2)

    args = parser.parse_args()
    
    # Tạo tên file output tự động nếu người dùng không cung cấp
    out_path = args.output
    if not out_path:
        parts = args.input.rsplit('.', 1)
        if len(parts) == 2:
            out_path = f"{parts[0]}_emails.{parts[1]}"
        else:
            out_path = args.input + "_emails.xlsx"

    process_excel(args.input, out_path, args.col, args.target)
