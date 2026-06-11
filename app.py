"""
app.py — Flask server chính
Chạy web:     python app.py
Chạy desktop: python desktop.py
"""

import io
import logging
import os
import re
import sys
import subprocess
import threading
import time
import concurrent.futures
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
import yt_dlp

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20 MB

# ── App State & Stats ────────────────────────────────────────────────────────
download_state = {
    "stop_requested": False,
    "is_downloading": False
}
ALLOWED_URLS = set()

# Global state cho email process
email_task_state = {
    "status": "idle", # idle, running, stopped, completed, error
    "progress": 0,
    "total_rows": 0,
    "processed_rows": 0,
    "emails_found": 0,
    "output_file": "",
    "error_msg": "",
    "sheet_stats": {} # Để vẽ chart: {"Sheet1": 10, "Sheet2": 5}
}

app_stats = {
    "yt_total_links": 0,
    "yt_valid_links": 0,
    "yt_downloaded": 0,
    "email_extracted_total": 0,
}

class DownloadCancelled(Exception):
    pass

def yt_progress_hook(d):
    if download_state["stop_requested"]:
        raise DownloadCancelled("Download stopped by user")

# Tắt cảnh báo "development server" và các log mặc định của Flask
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)


# ── Helpers (YouTube) ────────────────────────────────────────────────────────

def is_youtube(url: str) -> bool:
    return "youtube.com" in url.lower() or "youtu.be" in url.lower()


def extract_links_from_workbook(wb: openpyxl.Workbook) -> list[dict]:
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip()
                hyperlink = getattr(cell, "hyperlink", None)
                href = hyperlink.target if hyperlink and hyperlink.target else ""

                candidates = {val, href}
                for candidate in candidates:
                    if candidate and ("youtube" in candidate or "youtu.be" in candidate):
                        results.append({
                            "sheet": sheet_name,
                            "row": cell.row,
                            "col": cell.column,
                            "col_letter": cell.column_letter,
                            "title": val if val != candidate else "",
                            "url": candidate,
                            "valid": is_youtube(candidate),
                        })
                        break
    return results

# ── Helpers (Email) ──────────────────────────────────────────────────────────

EMAIL_REGEX = re.compile(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+')

def sort_emails(emails):
    unique_emails = list(set([e.lower() for e in emails]))
    unique_emails.sort(key=lambda e: (e.split('@')[1], e.split('@')[0]))
    return unique_emails

def process_email_thread_func(input_path, output_path, source_col, target_col):
    global email_task_state, app_stats
    email_task_state["status"] = "running"
    email_task_state["progress"] = 0
    email_task_state["processed_rows"] = 0
    email_task_state["emails_found"] = 0
    email_task_state["error_msg"] = ""
    email_task_state["sheet_stats"] = {}
    email_task_state["output_file"] = ""

    try:
        wb = openpyxl.load_workbook(input_path)
        
        # Calculate total rows first
        total_rows = 0
        for sheet_name in wb.sheetnames:
            total_rows += wb[sheet_name].max_row - 1 # exclude header
        email_task_state["total_rows"] = max(total_rows, 1)

        processed = 0
        total_found = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            email_task_state["sheet_stats"][sheet_name] = 0
            
            if ws.cell(row=1, column=target_col).value is None:
                ws.cell(row=1, column=target_col, value="Emails Được Trích Xuất")

            for row in range(2, ws.max_row + 1):
                if email_task_state["status"] == "stopped":
                    break

                cell_value = ws.cell(row=row, column=source_col).value
                if cell_value and isinstance(cell_value, str):
                    found_emails = EMAIL_REGEX.findall(cell_value)
                    if found_emails:
                        sorted_emails = sort_emails(found_emails)
                        ws.cell(row=row, column=target_col, value=", ".join(sorted_emails))
                        
                        count = len(sorted_emails)
                        total_found += count
                        email_task_state["sheet_stats"][sheet_name] += count
                        
                processed += 1
                email_task_state["processed_rows"] = processed
                email_task_state["emails_found"] = total_found
                email_task_state["progress"] = int((processed / email_task_state["total_rows"]) * 100)
            
            if email_task_state["status"] == "stopped":
                break

        wb.save(output_path)
        email_task_state["output_file"] = output_path
        app_stats["email_extracted_total"] += total_found

        if email_task_state["status"] != "stopped":
            email_task_state["status"] = "completed"
            email_task_state["progress"] = 100

    except Exception as e:
        email_task_state["status"] = "error"
        email_task_state["error_msg"] = str(e)
    finally:
        # Cleanup input temp file if needed (we'll keep it simple for now)
        pass

# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

# --- YT Routes ---

@app.route("/upload", methods=["POST"])
def upload():
    global ALLOWED_URLS, app_stats
    if "file" not in request.files:
        return jsonify({"error": "Không có file được gửi lên"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Tên file trống"}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xlsm", "xltx", "xltm"):
        return jsonify({"error": f"Định dạng .{ext} không hỗ trợ. Dùng .xlsx"}), 400

    try:
        data = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        links = extract_links_from_workbook(wb)
        
        def get_meta(l):
            if not l["valid"]:
                return None
            try:
                ydl_opts = {'quiet': True, 'no_warnings': True, 'extract_flat': True}
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    info = ydl.extract_info(l["url"], download=False)
                    
                    channel_name = info.get("uploader") or info.get("channel") or info.get("title") or "Không xác định"
                    channel_url = info.get("uploader_url") or info.get("channel_url")
                    
                    if not channel_url and 'entries' in info:
                        channel_url = l["url"]
                        
                    if channel_url:
                        if not channel_url.startswith("http"):
                            channel_url = f"https://www.youtube.com{channel_url}" if channel_url.startswith("/") else f"https://www.youtube.com/{channel_url}"
                        return {
                            "channel": channel_name,
                            "url": channel_url
                        }
                    return None
            except Exception as e:
                print(f"Error extracting {l['url']}: {e}")
                return None

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            meta_results = list(executor.map(get_meta, links))
            
        unique_channels = {}
        channel_counts = {}
        for m in meta_results:
            if m and m["url"]:
                unique_channels[m["url"]] = m["channel"]
                channel_counts[m["channel"]] = channel_counts.get(m["channel"], 0) + 1

        channels_list = [{"channel": v, "url": k} for k, v in unique_channels.items()]
        
        app_stats["yt_total_links"] += len(links)
        app_stats["yt_valid_links"] += len([l for l in links if l["valid"]])
        
        if "yt_channel_stats" not in app_stats:
            app_stats["yt_channel_stats"] = {}
        for ch, count in channel_counts.items():
            app_stats["yt_channel_stats"][ch] = app_stats["yt_channel_stats"].get(ch, 0) + count

        return jsonify({
            "filename": file.filename,
            "sheets": wb.sheetnames,
            "channels": channels_list,
            "total": len(links),
            "total_channels": len(channels_list)
        })
    except Exception as exc:
        return jsonify({"error": f"Lỗi đọc file: {exc}"}), 500


@app.route("/download_all", methods=["POST"])
def download_all():
    global app_stats
    data = request.json or {}
    urls = data.get("urls", [])
    if not urls:
        return jsonify({"error": "Không có URL nào để tải"}), 400
        
    valid_urls = [u for u in urls if u in ALLOWED_URLS]
    if not valid_urls:
        return jsonify({"error": "Tất cả link đều không hợp lệ hoặc không có trong file Excel!"}), 403

    download_state["stop_requested"] = False
    download_state["is_downloading"] = True

    download_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads")
    os.makedirs(download_dir, exist_ok=True)

    ydl_opts = {
        'outtmpl': os.path.join(download_dir, '%(title)s.%(ext)s'),
        'format': 'best',
        'quiet': False,
        'progress_hooks': [yt_progress_hook],
    }

    success_count = 0
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            for url in valid_urls:
                if download_state["stop_requested"]:
                    break
                try:
                    ydl.download([url])
                    success_count += 1
                    app_stats["yt_downloaded"] += 1
                except DownloadCancelled:
                    break
                except Exception as e:
                    print(f"Lỗi tải {url}: {e}")
                    continue
        
        if download_state["stop_requested"]:
            return jsonify({"message": f"Đã dừng tải. Thành công {success_count}/{len(valid_urls)} video!", "folder": download_dir})
        return jsonify({"message": f"Tải thành công {success_count}/{len(valid_urls)} video!", "folder": download_dir})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    finally:
        download_state["is_downloading"] = False
        download_state["stop_requested"] = False


@app.route("/stop_download", methods=["POST"])
def stop_download():
    if download_state["is_downloading"]:
        download_state["stop_requested"] = True
        return jsonify({"message": "Đang dừng quá trình tải..."})
    return jsonify({"message": "Không có tiến trình tải nào đang chạy."})


@app.route("/open_folder", methods=["POST"])
def open_folder():
    download_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads")
    os.makedirs(download_dir, exist_ok=True)
    try:
        if sys.platform == "win32":
            os.startfile(download_dir)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", download_dir])
        else:
            subprocess.Popen(["xdg-open", download_dir])
        return jsonify({"message": "Đã mở thư mục tải về"})
    except Exception as e:
        return jsonify({"error": f"Không thể mở thư mục: {e}"}), 500

@app.route("/fetch_channel", methods=["POST"])
def fetch_channel():
    global ALLOWED_URLS
    data = request.json or {}
    url = data.get("url")
    if not url:
        return jsonify({"error": "Vui lòng nhập link Channel hoặc Playlist"}), 400

    # extract_flat=True để lấy danh sách nhanh mà không tải chi tiết từng video
    ydl_opts = {
        'extract_flat': True,
        'quiet': True,
        'no_warnings': True,
    }
    
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)
            
            videos = []
            
            def extract_entries(entries_list):
                for entry in entries_list:
                    if not entry: continue
                    _type = entry.get('_type')
                    if _type == 'playlist' or 'entries' in entry:
                        tab_url = entry.get('url') or entry.get('webpage_url')
                        if tab_url:
                            try:
                                tab_info = ydl.extract_info(tab_url, download=False)
                                if 'entries' in tab_info:
                                    extract_entries(tab_info['entries'])
                            except:
                                pass
                        elif 'entries' in entry:
                            extract_entries(entry['entries'])
                    elif _type == 'url' or entry.get('url'):
                        v_url = entry.get('url') or entry.get('webpage_url')
                        if v_url:
                            if not v_url.startswith("http"):
                                v_url = f"https://www.youtube.com/watch?v={v_url}"
                            v_type = 'short' if '/shorts/' in v_url else 'video'
                            videos.append({
                                "title": entry.get('title', 'Không có tiêu đề'),
                                "url": v_url,
                                "type": v_type
                            })

            if 'entries' in info:
                extract_entries(info['entries'])
            else:
                v_url = info.get("url") or info.get("webpage_url")
                if v_url:
                    if not v_url.startswith("http"):
                        v_url = f"https://www.youtube.com/watch?v={v_url}"
                    v_type = 'short' if '/shorts/' in v_url else 'video'
                    videos.append({
                        "title": info.get("title", "Không có tiêu đề"),
                        "url": v_url,
                        "type": v_type
                    })
            
            # Thêm vào danh sách cho phép tải
            for v in videos:
                ALLOWED_URLS.add(v["url"])
                
            return jsonify({
                "channel_title": info.get("title", "Channel/Playlist"),
                "videos": videos,
                "total": len(videos)
            })
    except Exception as e:
        return jsonify({"error": f"Lỗi khi quét link: {str(e)}"}), 500

# --- Email Routes ---

@app.route("/upload_email_file", methods=["POST"])
def upload_email_file():
    if "file" not in request.files:
        return jsonify({"error": "Không có file được gửi lên"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Tên file trống"}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xlsm"):
        return jsonify({"error": f"Định dạng .{ext} không hỗ trợ. Dùng .xlsx"}), 400
        
    source_col = int(request.form.get("source_col", 1))
    target_col = int(request.form.get("target_col", 2))

    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads")
    os.makedirs(upload_dir, exist_ok=True)
    
    input_path = os.path.join(upload_dir, "temp_email_input.xlsx")
    output_path = os.path.join(upload_dir, f"{file.filename.rsplit('.', 1)[0]}_emails.xlsx")
    
    file.save(input_path)

    global email_task_state
    if email_task_state["status"] == "running":
        return jsonify({"error": "Một tiến trình khác đang chạy"}), 400

    # Khởi chạy thread
    thread = threading.Thread(target=process_email_thread_func, args=(input_path, output_path, source_col, target_col))
    thread.start()

    return jsonify({"message": "Đã bắt đầu xử lý email", "filename": file.filename})

@app.route("/stop_email_process", methods=["POST"])
def stop_email_process():
    global email_task_state
    if email_task_state["status"] == "running":
        email_task_state["status"] = "stopped"
        return jsonify({"message": "Đang dừng quá trình xử lý..."})
    return jsonify({"message": "Không có tiến trình nào đang chạy."})

@app.route("/email_progress", methods=["GET"])
def email_progress():
    global email_task_state
    return jsonify(email_task_state)
    
@app.route("/download_email_result", methods=["GET"])
def download_email_result():
    global email_task_state
    file_path = email_task_state.get("output_file")
    if file_path and os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return jsonify({"error": "File không tồn tại"}), 404

# --- Stats Route ---

@app.route("/stats", methods=["GET"])
def get_stats():
    global app_stats, email_task_state
    return jsonify({
        "app_stats": app_stats,
        "email_sheet_stats": email_task_state.get("sheet_stats", {})
    })

if __name__ == "__main__":
    print("🌐  Web đang chạy tại: http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
